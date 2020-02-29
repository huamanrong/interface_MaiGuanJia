from common import project_path
from openpyxl import load_workbook
import copy


class DoExcel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def read_data(self, mode, case_list=None):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        test_data = []  # 存储所有行的数据
        for i in range(2, sheet.max_row+1):
            sub_data = {}
            sub_data['case_id'] = sheet.cell(i, 1).value
            sub_data['method'] = sheet.cell(i, 3).value
            sub_data['url'] = sheet.cell(i, 4).value
            sub_data['json'] = sheet.cell(i, 5).value
            sub_data['param'] = sheet.cell(i, 6).value
            sub_data['correlation'] = sheet.cell(i, 7).value
            sub_data['expect_result'] = sheet.cell(i, 8).value
            sub_data['check_sql'] = sheet.cell(i, 11).value
            sub_data['preposition_sql'] = sheet.cell(i, 14).value
            sub_data['post_sql'] = sheet.cell(i, 15).value
            DoExcel.csv_data_set_config(test_data, sub_data, eval(sub_data['param']), project_path.csv_param_path)
            # test_data.append(sub_data)  # 所有的数据都存在test_data里面
        if mode == '1':
            final_data = test_data  # 返回所有的数据
        else:
            final_data = []
            for item in test_data:
                if item['case_id'] in case_list:
                    final_data.append(item)
        return final_data

    def write_data(self, row, mode, actually, result):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        if str(mode).lower() == 'request':
            sheet.cell(row, 9).value = actually
            sheet.cell(row, 10).value = result
        elif str(mode).lower() == 'sql_request':
            sheet.cell(row, 12).value = actually
            sheet.cell(row, 13).value = result
        wb.save(self.file_path)

    # 实现csv_data_set_config功能，暂时只支持单个接口使用csv_data_set_config功能
    @staticmethod
    def csv_data_set_config(total_data, row_param, param, file_path):
        if 'CSV_Data' in param.keys():
            print('已进入')
            csv_data = param['CSV_Data']
            param.pop('CSV_Data')
            with open(file_path, 'r', encoding='utf-8') as file:
                for value in file:
                    value_list = value.strip('\n').split(',')
                    for index, csv_value in enumerate(csv_data):
                        param[csv_value] = value_list[index]
                        row_param['param'] = str(param)
                    new_row_param = copy.deepcopy(row_param)
                    total_data.append(new_row_param)
        else:
            total_data.append(row_param)


if __name__ == '__main__':
    data = DoExcel(r'C:\Users\10336\Desktop\git_xinai\interface_LaoYiBao\test_data\test_interface.xlsx', '暂不运行用例').read_data('0', [1])
